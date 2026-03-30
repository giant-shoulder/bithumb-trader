"""
빗썸 자동매매 모니터링 대시보드
"""
import subprocess
import os
import csv
import glob
import io
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, date
from functools import wraps
from flask import Flask, render_template, Response, request, session, redirect, url_for, jsonify

app = Flask(__name__)
app.secret_key = os.environ.get('DASHBOARD_SECRET', 'bithumb-dashboard-secret')
DASHBOARD_PASSWORD = os.environ.get('DASHBOARD_PASSWORD', 'admin1234')


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        if request.form.get('password') == DASHBOARD_PASSWORD:
            session['logged_in'] = True
            return redirect(url_for('index'))
        return render_template('login.html', error='비밀번호가 틀렸습니다')
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/')
@login_required
def index():
    return render_template('index.html')


@app.route('/api/status')
@login_required
def status():
    result = subprocess.run(
        ['sudo', 'systemctl', 'is-active', 'bithumb-trader'],
        capture_output=True, text=True
    )
    return jsonify({'status': result.stdout.strip()})


@app.route('/api/control/<action>', methods=['POST'])
@login_required
def control(action):
    if action not in ['start', 'stop', 'restart']:
        return jsonify({'error': 'invalid action'}), 400
    subprocess.run(['sudo', 'systemctl', action, 'bithumb-trader'])
    return jsonify({'ok': True})


@app.route('/api/logs')
@login_required
def logs():
    def generate():
        process = subprocess.Popen(
            ['sudo', 'journalctl', '-u', 'bithumb-trader', '-f', '-n', '200', '--no-pager', '--output=short'],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True
        )
        for line in process.stdout:
            yield f"data: {line.rstrip()}\n\n"
    return Response(generate(), mimetype='text/event-stream',
                    headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'})


@app.route('/review')
@app.route('/review/<target_date>')
@login_required
def review(target_date=None):
    if target_date is None:
        target_date = date.today().strftime('%Y-%m-%d')
    return render_template('review.html', target_date=target_date)


@app.route('/api/review/<target_date>')
@login_required
def review_data(target_date):
    base_dir = '/home/ubuntu/bithumb-trader'
    ym = target_date[:7].replace('-', '')

    # 매매 기록 읽기
    trades = []
    trade_file = f"{base_dir}/trade_history_{ym}.csv"
    if os.path.exists(trade_file):
        with open(trade_file, encoding='utf-8') as f:
            reader = csv.DictReader(f, restkey=None)
            for row in reader:
                if row['시간'].startswith(target_date):
                    row.pop(None, None)
                    trades.append(row)

    # 탈락 기록 읽기
    rejects = []
    reject_file = f"{base_dir}/reject_history_{ym}.csv"
    if os.path.exists(reject_file):
        with open(reject_file, encoding='utf-8') as f:
            reader = csv.DictReader(f, restkey=None)
            for row in reader:
                if row['시간'].startswith(target_date):
                    row.pop(None, None)
                    rejects.append(row)

    # 통계 계산
    buys = [t for t in trades if t['유형'] == '매수']
    sells = [t for t in trades if t['유형'] == '매도']
    wins = [t for t in sells if float(t['손익률']) > 0]
    losses = [t for t in sells if float(t['손익률']) <= 0]
    total_pnl = sum(float(t['손익률']) for t in sells)

    # 실제 손익(원) 계산: 매수-매도 매칭으로 실제 투자금액 기준 계산
    # 각 매도에 대해 직전 매수들을 찾아 실제 손익 = 매도금액 - 매수금액 합계
    buy_queue = {}  # coin → list of buy amounts (chronological)
    for t in trades:
        coin = t['코인']
        if t['유형'] == '매수':
            buy_queue.setdefault(coin, []).append(float(t['금액']))
        elif t['유형'] == '매도':
            invested = sum(buy_queue.pop(coin, []))
            t['_buy_amt'] = invested

    pnl_krw = 0.0
    total_invested = 0.0
    for t in sells:
        sell_amt = float(t['금액'])
        buy_amt = t.get('_buy_amt', 0.0)
        if buy_amt > 0:
            pnl_krw += sell_amt - buy_amt
            total_invested += buy_amt
        else:
            # 매수 기록 없으면 손익률로 역산
            pnl_pct = float(t['손익률'])
            derived_buy = sell_amt / (1 + pnl_pct / 100)
            pnl_krw += derived_buy * pnl_pct / 100
            total_invested += derived_buy

    portfolio_return_pct = pnl_krw / total_invested * 100 if total_invested else 0

    # 신호 출처별 통계
    sources = {}
    for t in sells:
        src = t.get('신호출처', 'momentum')
        if src not in sources:
            sources[src] = {'count': 0, 'wins': 0, 'pnl': 0.0}
        sources[src]['count'] += 1
        sources[src]['pnl'] += float(t['손익률'])
        if float(t['손익률']) > 0:
            sources[src]['wins'] += 1

    return jsonify({
        'date': target_date,
        'trades': trades,
        'rejects': rejects,
        'stats': {
            'buy_count': len(buys),
            'sell_count': len(sells),
            'win_count': len(wins),
            'loss_count': len(losses),
            'win_rate': len(wins) / len(sells) * 100 if sells else 0,
            'total_pnl_pct': portfolio_return_pct,
            'total_pnl_krw': pnl_krw,
            'total_pnl_pct_sum': total_pnl,
            'avg_win': sum(float(t['손익률']) for t in wins) / len(wins) if wins else 0,
            'avg_loss': sum(float(t['손익률']) for t in losses) / len(losses) if losses else 0,
        },
        'sources': sources,
    })


@app.route('/api/review/<target_date>/export/xml')
@login_required
def export_xml(target_date):
    base_dir = '/home/ubuntu/bithumb-trader'
    ym = target_date[:7].replace('-', '')

    trades = []
    trade_file = f"{base_dir}/trade_history_{ym}.csv"
    if os.path.exists(trade_file):
        with open(trade_file, encoding='utf-8') as f:
            reader = csv.DictReader(f, restkey=None)
            for row in reader:
                if row['시간'].startswith(target_date):
                    row.pop(None, None)
                    trades.append(row)

    rejects = []
    reject_file = f"{base_dir}/reject_history_{ym}.csv"
    if os.path.exists(reject_file):
        with open(reject_file, encoding='utf-8') as f:
            reader = csv.DictReader(f, restkey=None)
            for row in reader:
                if row['시간'].startswith(target_date):
                    row.pop(None, None)
                    rejects.append(row)

    root = ET.Element('매매회고', 날짜=target_date)

    trades_el = ET.SubElement(root, '매매내역')
    for t in trades:
        trade_el = ET.SubElement(trades_el, '거래')
        for k, v in t.items():
            if not k.startswith('_'):
                ET.SubElement(trade_el, k).text = str(v or '')

    rejects_el = ET.SubElement(root, '탈락내역')
    for r in rejects:
        reject_el = ET.SubElement(rejects_el, '탈락')
        for k, v in r.items():
            if not k.startswith('_'):
                ET.SubElement(reject_el, k).text = str(v or '')

    xml_bytes = ET.tostring(root, encoding='utf-8', xml_declaration=True)
    return Response(
        xml_bytes,
        mimetype='application/xml',
        headers={'Content-Disposition': f'attachment; filename="trades_{target_date}.xml"'}
    )


@app.route('/api/review/<target_date>/export/excel')
@login_required
def export_excel(target_date):
    base_dir = '/home/ubuntu/bithumb-trader'
    ym = target_date[:7].replace('-', '')

    trades = []
    trade_file = f"{base_dir}/trade_history_{ym}.csv"
    if os.path.exists(trade_file):
        with open(trade_file, encoding='utf-8') as f:
            reader = csv.DictReader(f, restkey=None)
            for row in reader:
                if row['시간'].startswith(target_date):
                    row.pop(None, None)
                    trades.append(row)

    rejects = []
    reject_file = f"{base_dir}/reject_history_{ym}.csv"
    if os.path.exists(reject_file):
        with open(reject_file, encoding='utf-8') as f:
            reader = csv.DictReader(f, restkey=None)
            for row in reader:
                if row['시간'].startswith(target_date):
                    row.pop(None, None)
                    rejects.append(row)

    wb = openpyxl.Workbook()

    # ── 매매내역 시트 ──
    ws = wb.active
    ws.title = '매매내역'

    header_fill = PatternFill('solid', fgColor='1F2937')
    header_font = Font(color='79C0FF', bold=True)
    buy_fill   = PatternFill('solid', fgColor='1A4731')
    sell_fill  = PatternFill('solid', fgColor='3D1A1A')

    if trades:
        cols = [k for k in trades[0].keys() if not k.startswith('_')]
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for ri, t in enumerate(trades, 2):
            is_buy = t.get('유형') == '매수'
            row_fill = buy_fill if is_buy else sell_fill
            for ci, col in enumerate(cols, 1):
                val = t.get(col, '')
                try:
                    val = float(val)
                except (ValueError, TypeError):
                    pass
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.fill = row_fill

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 16

    # ── 탈락내역 시트 ──
    ws2 = wb.create_sheet('탈락내역')
    if rejects:
        cols2 = [k for k in rejects[0].keys() if not k.startswith('_')]
        for ci, col in enumerate(cols2, 1):
            cell = ws2.cell(row=1, column=ci, value=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for ri, r in enumerate(rejects, 2):
            for ci, col in enumerate(cols2, 1):
                val = r.get(col, '')
                try:
                    val = float(val)
                except (ValueError, TypeError):
                    pass
                ws2.cell(row=ri, column=ci, value=val)
        for col in ws2.columns:
            ws2.column_dimensions[col[0].column_letter].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="trades_{target_date}.xlsx"'}
    )


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=False)
